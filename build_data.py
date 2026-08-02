#!/usr/bin/env python3
"""
Build data/collection.json for the Pokémon TCG Collection website
from the "PTCG Purchases Tracker.xlsx" spreadsheet.

Workflow for updates:
    1. Keep the xlsx updated.
    2. Run:  python3 build_data.py  [optional/path/to/tracker.xlsx]
    3. Commit & push  data/collection.json  (bump sw.js cache name).

Reads two sheets:
    "Card Purchases" -> one row per card line (name, qty, all-in cost per card)
    "Needs"          -> quantity + card name of cards still wanted

Cards are grouped by a normalised BASE NAME so that different printings /
variants of the same card sit together, while each distinct printed name
(including its [RH]/[PP]/[Holo]/[FA]/(pattern)/set-code tags) stays a
separate variant row.
"""
import json
import os
import re
import sys
from collections import OrderedDict

DEFAULT_XLSX = "/Users/vighnesh/Library/CloudStorage/OneDrive-Personal/01. Documents/05. Games & Books/TCG/PTCG Purchases Tracker.xlsx"
OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "collection.json")

# --- variant token expansion (display only; unknown tokens shown verbatim) ---
TOKEN_MAP = {
    "RH": "Reverse Holo",
    "PP": "Prize Pack",
    "Holo": "Holo",
    "FA": "Full Art",
    "PB": "Poké Ball",
}


def clean_ws(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip()


def parse_name(full: str):
    """Return (base_name, [variant_tags]) for a printed card name."""
    tags = []
    work = full

    # 1) parenthetical tags anywhere e.g. (Ascended Heroes), (Poke Ball), (Energy Symbol Pattern)
    for m in re.findall(r"\(([^)]*)\)", work):
        t = clean_ws(m)
        if t:
            tags.append(t)
    work = re.sub(r"\([^)]*\)", " ", work)

    # 2) set code e.g. " - 141/167", " - 025"
    setcodes = re.findall(r"-\s*(\d+(?:/\w+)?)\b", work)
    for sc in setcodes:
        tags.append("#" + sc)
    work = re.sub(r"-\s*\d+(?:/\w+)?\b", " ", work)

    # 3) bracket tags e.g. [RH], [PP], [Holo], [Corbeau]
    for m in re.findall(r"\[([^\]]*)\]", work):
        t = clean_ws(m)
        if t:
            tags.append(t)
    work = re.sub(r"\[[^\]]*\]", " ", work)

    # 4) trailing bare " FA" (Full Art) e.g. "Oricorio ex FA", "Tapu Lele GX FA"
    if re.search(r"\bFA\s*$", work):
        tags.append("FA")
        work = re.sub(r"\bFA\s*$", " ", work)

    base = clean_ws(work)
    if not base:  # safety: never blank
        base = clean_ws(full)
    return base, tags


def variant_label(tags):
    """Human-readable variant chip. Empty tags -> 'Standard'."""
    if not tags:
        return "Standard"
    out = []
    for t in tags:
        if t.startswith("#"):
            out.append("No. " + t[1:])
        else:
            out.append(TOKEN_MAP.get(t, t))
    return " · ".join(out)


def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(xlsx):
        sys.exit(f"ERROR: spreadsheet not found: {xlsx}")
    try:
        import openpyxl
    except ImportError:
        sys.exit("ERROR: openpyxl not installed. Run: python3 -m pip install openpyxl --break-system-packages")

    wb = openpyxl.load_workbook(xlsx, data_only=True)

    # ---- Card Purchases -> cost per BASE card name (for avg-cost display) ----
    cp = wb["Card Purchases"]
    rows = list(cp.iter_rows(values_only=True))
    header = [str(h).strip() if h else "" for h in rows[0]]
    col = {name: i for i, name in enumerate(header)}
    ci_name, ci_qty, ci_total = col["Card Name"], col["Quantity"], col["Total Cost"]

    cost_by_base = {}   # base name -> {"spent": all-in $, "pqty": purchased qty}
    total_spent = 0.0
    purchased_cards = 0
    for r in rows[1:]:
        if not r[ci_name]:
            continue
        base, _tags = parse_name(clean_ws(str(r[ci_name])))
        qty = r[ci_qty] if isinstance(r[ci_qty], (int, float)) else 0
        total = r[ci_total] if isinstance(r[ci_total], (int, float)) else 0.0   # skip #DIV/0! etc.
        cb = cost_by_base.setdefault(base, {"spent": 0.0, "pqty": 0})
        cb["spent"] += total
        cb["pqty"] += qty
        total_spent += total
        purchased_cards += qty

    # ---- Collection sheet -> OWNED inventory (source of truth for what I have) ----
    # Columns: Card Name | Card Type | Set | Card Number | Version | Quantity
    colws = wb["Collection"]
    owned = OrderedDict()   # base name -> group
    for r in list(colws.iter_rows(min_col=1, max_col=6, values_only=True))[1:]:
        name, ctype, cset, cnum, version, qty = r
        if not name or not isinstance(qty, (int, float)):
            continue
        name = clean_ws(str(name))
        version = clean_ws(str(version)) if version else "Standard"
        qty = int(qty)
        vl = version.lower()
        jp = vl == "japanese"
        proxy = vl == "proxy"
        cset = clean_ws(str(cset)) if cset not in (None, "") else None
        cnum = str(cnum).strip() if cnum not in (None, "") else None
        g = owned.get(name)
        if g is None:
            g = owned[name] = {"base": name, "type": clean_ws(str(ctype)) if ctype else None,
                               "variants": [], "total": 0, "totalNonJp": 0}
        g["variants"].append({"version": version, "qty": qty, "jp": jp, "proxy": proxy, "set": cset, "num": cnum})
        g["total"] += qty
        if not jp:
            g["totalNonJp"] += qty

    # sort variants (Standard first, Japanese last) + attach cost from purchases
    def vkey(v):
        return (v["jp"], v["version"] != "Standard", v["version"].lower())
    group_list = []
    for g in owned.values():
        g["variants"].sort(key=vkey)
        cb = cost_by_base.get(g["base"])
        if cb and cb["pqty"]:
            g["spent"] = round(cb["spent"], 2)
            g["pqty"] = cb["pqty"]
            g["avgCost"] = round(cb["spent"] / cb["pqty"], 2)
        else:
            g["spent"] = None
            g["pqty"] = 0
            g["avgCost"] = None
        group_list.append(g)
    group_list.sort(key=lambda g: g["base"].lower())

    owned_copies = sum(g["total"] for g in group_list)
    owned_entries = sum(len(g["variants"]) for g in group_list)
    jp_copies = sum(v["qty"] for g in group_list for v in g["variants"] if v["jp"])
    proxy_copies = sum(v["qty"] for g in group_list for v in g["variants"] if v["proxy"])

    # ---- Needs ----
    needs = []
    if "Needs" in wb.sheetnames:
        for r in list(wb["Needs"].iter_rows(values_only=True))[1:]:
            if not r or not r[1]:
                continue
            needs.append({"qty": r[0] or 0, "name": clean_ws(str(r[1]))})
        needs.sort(key=lambda x: x["name"].lower())

    from datetime import datetime, timezone
    data = {
        "updated": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "currency": "NZD",
        "stats": {
            "ownedCopies": owned_copies,
            "ownedNames": len(group_list),
            "ownedEntries": owned_entries,
            "japaneseCopies": jp_copies,
            "proxyCopies": proxy_copies,
            "totalSpent": round(total_spent, 2),
            "purchasedCards": purchased_cards,
            "avgCostPerCard": round(total_spent / purchased_cards, 2) if purchased_cards else 0.0,
            "needsCount": sum(n["qty"] for n in needs),
        },
        "groups": group_list,
        "needs": needs,
    }

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)

    print(f"Wrote {OUT_PATH}")
    print(f"  OWNED: {owned_copies} copies | {len(group_list)} names | {owned_entries} versions | {jp_copies} Japanese")
    print(f"  SPENT: ${total_spent:.2f} {data['currency']} over {purchased_cards} purchased cards")
    print(f"  needs: {len(needs)} lines / {data['stats']['needsCount']} cards")


if __name__ == "__main__":
    main()
